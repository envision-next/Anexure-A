"""
Loads people + their organizations from the first .xlsx in the folder.

The Excel is expected to have ONE ROW PER (person, organization) link. Column
headers are matched by keyword (case-insensitive), so the exact wording can vary
-- e.g. "RERA Registration Number", "RERA No.", "Rera number" all match.

If no .xlsx is found, built-in SAMPLE_PEOPLE (from the original template) is used
so the app is usable immediately.

Each person is a dict:
    {
      "name": "Lalji Bhachu Patni",
      "din":  "09403966",
      "orgs": [ {name, address, rera, completion, complaints, warrants, revoked}, ... ]
    }
"""

import os
import glob

import openpyxl


# ---- Column detection ------------------------------------------------------
# Each field maps to a list of keyword groups. A header matches a field if it
# contains ALL keywords of ANY one group. Order matters: more specific first.
COLUMN_KEYWORDS = {
    "name":       [["director", "name"], ["person"], ["member", "name"], ["name of"]],
    "din":        [["din"], ["dpin"]],
    "org":        [["organi"], ["firm"], ["company"]],
    "address":    [["address"]],
    "rera":       [["rera"], ["registration", "number"], ["reg", "no"]],
    "completion": [["completion"], ["proposed", "date"], ["date", "complet"]],
    "complaints": [["complaint"]],
    "warrants":   [["warrant"]],
    "revoked":    [["revoke"]],
}

# Fields where a plain "name" header should NOT be treated as the person name
# (organization name also contains "name").
_ORG_HINTS = ("organi", "firm", "company", "project")


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _match_columns(headers):
    """Return {field: column_index} for detected headers."""
    norm = [_norm(h) for h in headers]
    mapping = {}
    for field, groups in COLUMN_KEYWORDS.items():
        for ci, h in enumerate(norm):
            if not h or ci in mapping.values():
                continue
            # For person name, skip headers that look like an org/project name.
            if field == "name" and any(hint in h for hint in _ORG_HINTS):
                continue
            if any(all(kw in h for kw in group) for group in groups):
                mapping[field] = ci
                break
    return mapping


# ---- Sample fallback -------------------------------------------------------
SAMPLE_PEOPLE = [
    {
        "name": "Lalji Bhachu Patni",
        "din": "09403966",
        "orgs": [
            {
                "name": "Shreeji Lifespaces Buildcon",
                "address": "Office No. B-103, The Great Eastern Summit, "
                           "Plot No -66 Sector -15, CBD Belapur, Thane - 400614",
                "rera": "P52000046569",
                "completion": "31/12/2025",
                "complaints": "No",
                "warrants": "No",
                "revoked": "No",
            },
            {
                "name": "Today Royal Housing",
                "address": "1201, Siddhi Grandeur, Plot No.84, Sector- 19, "
                           "Kharghar, Navi Mumbai - 410210",
                "rera": "P52000080085",
                "completion": "30/06/2029",
                "complaints": "No",
                "warrants": "No",
                "revoked": "No",
            },
        ],
    }
]


def find_excel(base_dir):
    """First real .xlsx in the folder (ignores Word temp files like ~$...)."""
    for path in sorted(glob.glob(os.path.join(base_dir, "*.xlsx"))):
        if not os.path.basename(path).startswith("~$"):
            return path
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    # Dates -> dd/mm/yyyy
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_people(base_dir):
    """Build one searchable entity per 'Name of Organisation', each carrying all
    of that organisation's registered projects (one project = one Annexure row).

    The sheet has no person/DIN column, so the searched name IS the organisation
    (promoter). DIN/DPIN is entered by hand in the UI. Warrants / Revoked are not
    in the sheet, so they default to "No"; Complaint Number defaults to "No" when
    the cell is blank.
    """
    path = find_excel(base_dir)
    if not path:
        return [dict(p) for p in SAMPLE_PEOPLE]

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = next(it, None)
    if not headers:
        return [dict(p) for p in SAMPLE_PEOPLE]

    cols = _match_columns(headers)
    org_i = cols.get("org")
    if org_i is None:               # can't group without an organisation column
        return [dict(p) for p in SAMPLE_PEOPLE]

    entities = {}
    order = []
    for raw in it:
        if raw is None:
            continue
        org_name = _cell(raw, org_i)
        if not org_name:
            continue
        if org_name not in entities:
            entities[org_name] = {"name": org_name, "din": "", "orgs": []}
            order.append(org_name)
        entities[org_name]["orgs"].append({
            "name": org_name,
            "address": _cell(raw, cols.get("address")),
            "rera": _cell(raw, cols.get("rera")),
            "completion": _cell(raw, cols.get("completion")),
            "complaints": _cell(raw, cols.get("complaints")) or "No",
            "warrants": "No",
            "revoked": "No",
        })

    wb.close()
    result = [entities[k] for k in order]
    return result or [dict(p) for p in SAMPLE_PEOPLE]


def describe_source(base_dir):
    path = find_excel(base_dir)
    if not path:
        return {
            "using_excel": False,
            "label": "Sample data (no Excel found yet)",
            "file": None,
            "columns": {},
        }
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    headers = next(wb.active.iter_rows(values_only=True), ())
    wb.close()
    cols = _match_columns(headers)
    detected = {field: (headers[i] if i < len(headers) else None)
                for field, i in cols.items()}
    return {
        "using_excel": True,
        "label": os.path.basename(path),
        "file": os.path.basename(path),
        "columns": detected,
    }
